from datetime import timedelta
from pathlib import Path
from typing import Union

import numpy as np
import pandas as pd
import regex
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from typing import List
from .timeformat import dateFromText
from typing import Dict

def guardarExcel(
    df: pd.DataFrame,
    fname: Union[Path, str],
    sheet_name: str = "Sheet1",
    append_sheet: bool = True,
):
    """
    Guarda un DataFrame en un documento excel creando una tabla con headers

    append_sheet: bool
        - True:   mantiene la información anterior y escribe una nueva hoja
        - False:  sobrescribe el fichero
    """
    fname = Path(fname)
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    if fname.exists() and append_sheet:
        mode = "a"
    else:
        mode = "w"
    sheet_name = regex.sub(r"(\W|_)", "", sheet_name)
    
    if mode == "a":
        writer = pd.ExcelWriter(
            fname,
            mode="a",
            engine="openpyxl",
            if_sheet_exists="replace"
        )
    else:
        writer = pd.ExcelWriter(
            fname,
            mode="w",
            engine="openpyxl"
        )

    with writer:

        df.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
        wb = writer.book
        ws = wb[sheet_name]
        tab = Table(
            displayName=f"{sheet_name}_",
            name=sheet_name,
            ref=f"A1:{get_column_letter(df.shape[1])}{df.shape[0]+1}",
            tableStyleInfo=style,
        )
        ws.add_table(tab)

        column_widths = [
            max([len(c)] + df[c].fillna("").astype(str).apply(len).tolist())
            for c in df.columns
        ]
        for j, column_width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = min(
                [60, column_width + 4]
            )
        for row in ws:
            for cell in row:
                cell.alignment = Alignment(wrapText=True)


def guardarExcelMulti(
    dfs: Dict[str, pd.DataFrame],
    fname: Union[Path, str],
):
    """
    Guarda un DataFrames en un documento excel creando una tabla con headers.

    dfs: dict(str:DataFrame)
        - key: nombre de la hoja
        - value: datos
    """
    fname = Path(fname)
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    with pd.ExcelWriter(fname, mode="w", engine="openpyxl") as writer:
        for i, (sheet_name, df) in enumerate(dfs.items(), 1):
            print(f"Guardando: {sheet_name} ({i}/{len(dfs)})")
            df.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
            wb = writer.book
            ws = wb[sheet_name]
            tab = Table(
                displayName=f"{sheet_name}_",
                name=sheet_name,
                ref=f"A1:{get_column_letter(df.shape[1])}{df.shape[0]+1}",
                tableStyleInfo=style,
            )
            ws.add_table(tab)

            column_widths = [
                max([len(c)] + df[c].fillna("").astype(str).apply(len).tolist())
                for c in df.columns
            ]
            for j, column_width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(j)].width = min(
                    [60, column_width + 4]
                )
            for row in ws:
                for cell in row:
                    cell.alignment = Alignment(wrapText=True)


def getFilesByDate(
    dir_files: Path,
    start_date: str = "2024-01-01",
    end_date: str = "2024-12-31",
    ftype: str = "*",
):
    """
    Devuelve la lista de ficheros con logs entre dos fechas (ambas incluidas)
    """
    start_date = pd.to_datetime(start_date).date()
    end_date = pd.to_datetime(end_date).date()
    valid_days: list[tuple[Path, pd.Timestamp]] = []
    # Comprobar qué ficheros cargar
    # Puede haber dos formatos de fecha (YYYY-MM-dd) o (dds)
    if ftype == "*" or ftype is None:
        files = list(dir_files.rglob("*.*"))
    else:
        files = list(dir_files.rglob(f"*.{ftype}"))
    for fname in files:
        if ".zip" in fname.name:
            continue
        stem = fname.stem
        full_day = dateFromText(stem)
        if full_day is None:
            continue
        full_day = full_day.split(" - ")
        if len(full_day) == 1:
            fd0 = pd.to_datetime(full_day[0]).date()
            if fd0 >= start_date and fd0 <= end_date:
                valid_days.append((fname, fd0))
        elif len(full_day) == 2:
            fd0 = pd.to_datetime(full_day[0]).date()
            fd1 = pd.to_datetime(full_day[1]).date()
            if (fd0 >= start_date and fd0 <= end_date) or (
                fd1 >= start_date and fd1 <= end_date
            ):
                valid_days.append((fname, fd0))
                if fd0 != fd1:
                    for d in range((fd1 - fd0).days):
                        fd = (pd.to_datetime(fd0) + timedelta(days=d + 1)).date()
                        valid_days.append((fname, fd))
        else:
            print("?")
            continue
        # if "-" in stem:
        #     full_day = fname.stem.split(" ", 1)[-1].split(" - ")
        #     fd0 = pd.to_datetime(full_day[0]).date()
        #     fd1 = pd.to_datetime(full_day[-1]).date()
        #     if (fd0 >= start_date and fd0 <= end_date) or (
        #         fd1 >= start_date and fd1 <= end_date
        #     ):
        #         valid_days.append((fname, fd0))
        #         if fd0 != fd1:
        #             for d in range((fd1 - fd0).days):
        #                 fd = (
        #                     pd.to_datetime(full_day[0]) + timedelta(days=d + 1)
        #                 ).date()
        #                 valid_days.append((fname, fd))
        # else:
        #     # d = stem.split(".")[-1]
        #     d = regex.split(r"[\s\._]+", stem)[-1]
        #     print(d)
        #     m = fname.parent.stem
        #     full_day = pd.to_datetime(f"{m}-{d}").date()
        #     if full_day >= start_date and full_day <= end_date:
        #         valid_days.append((fname, full_day))
    sorted_days = [d for d in sorted(valid_days, key=lambda x: x[1])]
    return sorted_days


def getFilesByWeek(
    dir_logs: Path,
    start_date: str = "2024-01-01",
    end_date: str = "2024-12-31",
    start_of_week: int = 4,
):
    """
    Devuelve la lista de ficheros con logs entre dos fechas (ambas incluidas).
    Devuelve ÚNICAMENTE semanas enteras empezando desde el día de la semana elegido:
        - 0: lunes
        - 1: martes
        - ...
        - 4: viernes
    """
    days = getFilesByDate(dir_logs, start_date, end_date)
    fnames, full_days = list(zip(*days))
    days_of_week = np.array([d.day_of_week for d in full_days])
    w_split = np.where(days_of_week == start_of_week)[0]
    for dw, w, fd in zip(
        np.split(days_of_week, w_split),
        np.split(fnames, w_split),
        np.split(full_days, w_split),
    ):
        if np.in1d(range(7), dw).all():
            yield list(zip(w, fd))
